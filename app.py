
#!/usr/bin/env python3
"""
Enhanced Modern Web Scraper & Curl Platform (Flask + Bootstrap)
- Improvements added:
  * Robust request handling (headers dict, binary-safe decoding)
  * Auto-find contact page + email extraction (single-click)
  * Safer global storage for last results + proper use of `global`
  * Better error messages and exception handling
  * Email / phone / url regex presets
  * Improved download handling for curl/json results
  * Enhanced UI: beautiful borders, vibrant buttons, responsive design
  * Mode (curl/scrape) only changes on mode button click
  * Three themes: light (gradient bg, black text), dark (black bg, white text), dark-alt (white bg, black text)

Keep this file as `enhanced_scraper.py` and run with `python3 enhanced_scraper.py`.
"""

from flask import Flask, request, render_template_string, send_file, redirect, url_for, flash, session
import requests
from bs4 import BeautifulSoup
import pandas as pd
import io
import json
import re
from urllib.parse import urlparse, urljoin
from datetime import datetime
import openpyxl
import traceback

app = Flask(__name__)
app.secret_key = "change_this_secret_in_production_please"

# Enhanced UI template with beautiful styling
TEMPLATE = """
<!doctype html>
<html lang="en" data-bs-theme="{{ theme }}">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Enhanced Web Scraper & Curl</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
  <style>
    body { 
      min-height: 100vh; 
      font-family: 'Segoe UI', sans-serif;
      transition: all 0.3s ease;
    }
    body.theme-light {
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: #000000;
    }
    body.theme-dark {
      background: #1a1a1a;
      color: #ffffff;
    }
    body.theme-dark-alt {
      background: #ffffff;
      color: #000000;
    }
    body.theme-light .glass { 
      background: rgba(255,255,255,0.2); 
      backdrop-filter: blur(12px); 
      border-radius: 20px; 
      border: 2px solid rgba(0,123,255,0.5);
      padding: 20px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    }
    body.theme-dark .glass { 
      background: rgba(0,0,0,0.3); 
      backdrop-filter: blur(12px); 
      border-radius: 20px; 
      border: 2px solid rgba(255,255,255,0.3);
      padding: 20px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    }
    body.theme-dark-alt .glass { 
      background: rgba(255,255,255,0.2); 
      backdrop-filter: blur(12px); 
      border-radius: 20px; 
      border: 2px solid rgba(0,0,0,0.3);
      padding: 20px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    }
    body.theme-light .hero { 
      background: rgba(0,0,0,0.95); 
      border-radius: 20px; 
      box-shadow: 0 8px 32px rgba(0,0,0,0.15); 
      color: #ffffff; 
      padding: 30px;
    }
    body.theme-dark .hero { 
      background: rgba(255,255,255,0.95); 
      border-radius: 20px; 
      box-shadow: 0 8px 32px rgba(0,0,0,0.15); 
      color: #000000; 
      padding: 30px;
    }
    body.theme-dark-alt .hero { 
      background: rgba(0,0,0,0.95); 
      border-radius: 20px; 
      box-shadow: 0 8px 32px rgba(0,0,0,0.15); 
      color: #ffffff; 
      padding: 30px;
    }
    body.theme-light pre { 
      background: #f8f9fa; 
      border-radius: 10px; 
      padding: 20px; 
      overflow-x: auto; 
      max-height: 400px; 
      color: #000000;
      border: 1px solid #007bff;
    }
    body.theme-dark pre { 
      background: #2d2d2d; 
      border-radius: 10px; 
      padding: 20px; 
      overflow-x: auto; 
      max-height: 400px; 
      color: #ffffff;
      border: 1px solid #ffffff;
    }
    body.theme-dark-alt pre { 
      background: #f8f9fa; 
      border-radius: 10px; 
      padding: 20px; 
      overflow-x: auto; 
      max-height: 400px; 
      color: #000000;
      border: 1px solid #007bff;
    }
    body.theme-light .form-control, body.theme-light .form-select {
      border: 2px solid #007bff;
      border-radius: 10px;
      background: rgba(255,255,255,0.9);
      color: #000000;
      transition: border-color 0.3s ease;
    }
    body.theme-dark .form-control, body.theme-dark .form-select {
      border: 2px solid #ffffff;
      border-radius: 10px;
      background: rgba(255,255,255,0.1);
      color: #ffffff;
      transition: border-color 0.3s ease;
    }
    body.theme-dark-alt .form-control, body.theme-dark-alt .form-select {
      border: 2px solid #333333;
      border-radius: 10px;
      background: rgba(255,255,255,0.9);
      color: #000000;
      transition: border-color 0.3s ease;
    }
    body.theme-light .form-control:focus, body.theme-light .form-select:focus {
      border-color: #0056b3;
      box-shadow: 0 0 8px rgba(0,123,255,0.5);
    }
    body.theme-dark .form-control:focus, body.theme-dark .form-select:focus {
      border-color: #00d4ff;
      box-shadow: 0 0 8px rgba(0,212,255,0.5);
    }
    body.theme-dark-alt .form-control:focus, body.theme-dark-alt .form-select:focus {
      border-color: #555555;
      box-shadow: 0 0 8px rgba(0,123,255,0.5);
    }
    .btn-modern {
      border-radius: 50px;
      padding: 12px 30px;
      font-weight: 600;
      transition: all 0.3s ease;
      border: none;
    }
    .btn-primary {
      background: linear-gradient(45deg, #007bff, #00d4ff);
      color: #ffffff;
    }
    .btn-primary:hover {
      transform: scale(1.05);
      box-shadow: 0 5px 15px rgba(0,123,255,0.4);
    }
    body.theme-light .btn-outline-light {
      border: 2px solid #007bff;
      color: #007bff;
    }
    body.theme-dark .btn-outline-light {
      border: 2px solid #ffffff;
      color: #ffffff;
    }
    body.theme-dark-alt .btn-outline-light {
      border: 2px solid #333333;
      color: #333333;
    }
    body.theme-light .btn-outline-light:hover {
      background: rgba(0,123,255,0.2);
      transform: scale(1.05);
    }
    body.theme-dark .btn-outline-light:hover {
      background: rgba(255,255,255,0.2);
      transform: scale(1.05);
    }
    body.theme-dark-alt .btn-outline-light:hover {
      background: rgba(0,123,255,0.2);
      transform: scale(1.05);
    }
    .btn-success {
      background: linear-gradient(45deg, #28a745, #34c759);
      color: #ffffff;
    }
    .btn-success:hover {
      transform: scale(1.05);
      box-shadow: 0 5px 15px rgba(40,167,69,0.4);
    }
    body.theme-light .mode-toggle {
      cursor: pointer;
      transition: all 0.3s ease;
      border-radius: 50px;
      padding: 12px 20px;
      background: rgba(255,255,255,0.2);
      border: 2px solid #007bff;
      color: #000000;
    }
    body.theme-dark .mode-toggle {
      cursor: pointer;
      transition: all 0.3s ease;
      border-radius: 50px;
      padding: 12px 20px;
      background: rgba(255,255,255,0.1);
      border: 2px solid #ffffff;
      color: #ffffff;
    }
    body.theme-dark-alt .mode-toggle {
      cursor: pointer;
      transition: all 0.3s ease;
      border-radius: 50px;
      padding: 12px 20px;
      background: rgba(0,0,0,0.1);
      border: 2px solid #333333;
      color: #000000;
    }
    .mode-toggle.active {
      background: linear-gradient(45deg, #007bff, #00d4ff);
      color: #ffffff;
      transform: scale(1.05);
      border: none;
    }
    body.theme-light .mode-toggle:hover {
      background: rgba(0,123,255,0.3);
    }
    body.theme-dark .mode-toggle:hover {
      background: rgba(255,255,255,0.3);
    }
    body.theme-dark-alt .mode-toggle:hover {
      background: rgba(0,123,255,0.3);
    }
    body.theme-light .progress {
      height: 8px;
      background: rgba(0,0,0,0.1);
      border-radius: 10px;
    }
    body.theme-dark .progress {
      height: 8px;
      background: rgba(255,255,255,0.2);
      border-radius: 10px;
    }
    body.theme-dark-alt .progress {
      height: 8px;
      background: rgba(0,0,0,0.1);
      border-radius: 10px;
    }
    .progress-bar {
      background: linear-gradient(45deg, #007bff, #00d4ff);
    }
    body.theme-light .alert {
      border-radius: 10px;
      border: 2px solid #007bff;
    }
    body.theme-dark .alert {
      border-radius: 10px;
      border: 2px solid #ffffff;
    }
    body.theme-dark-alt .alert {
      border-radius: 10px;
      border: 2px solid #333333;
    }
    body.theme-light .list-group-item {
      background: rgba(255,255,255,0.9);
      border: 1px solid #007bff;
      color: #000000;
    }
    body.theme-dark .list-group-item {
      background: rgba(255,255,255,0.1);
      border: 1px solid #ffffff;
      color: #ffffff;
    }
    body.theme-dark-alt .list-group-item {
      background: rgba(255,255,255,0.9);
      border: 1px solid #007bff;
      color: #000000;
    }
    body.theme-light .text-muted {
      color: #ffffff !important;
    }
    body.theme-dark .text-muted {
      color: #cccccc !important;
    }
    body.theme-dark-alt .text-muted {
      color: #666666 !important;
    }
    @media (max-width: 576px) {
      .hero { padding: 20px; }
      .btn-modern { padding: 10px 20px; }
      .mode-toggle { padding: 10px 15px; }
    }
  </style>
</head>
<body class="d-flex align-items-center min-vh-100 py-3 theme-{{ theme }}">
  <div class="container">
    <div class="row justify-content-center">
      <div class="col-lg-10 col-md-12">
        <div class="d-flex justify-content-between align-items-center mb-4">
          <h1 class="display-5 fw-bold"><i class="fas fa-spider me-3"></i>Web Scraper & Curl</h1>
          <button class="btn btn-outline-light btn-modern" onclick="toggleDarkMode()"><i class="fas fa-moon"></i> Theme</button>
        </div>
        <div class="hero p-4 p-lg-5 mb-4">
          <p class="text-center text-muted mb-4 fs-5">Advanced fetch, scrape, export with ease.</p>

          <form method="post" action="{{ url_for('process') }}" id="scrapeForm">
            <input type="hidden" name="autofind" id="autofind" value="0">
            <input type="hidden" name="theme" id="theme" value="{{ theme }}">
            <input type="hidden" name="mode" id="mode" value="{{ request.form.get('mode', 'curl') }}">
            <div class="row mb-4">
              <div class="col-12">
                <label class="form-label fw-bold fs-6">Target URL</label>
                <div class="input-group">
                  <span class="input-group-text"><i class="fas fa-link"></i></span>
                  <input name="url" type="url" class="form-control" placeholder="https://example.com" required value="{{ request.form.get('url','') }}">
                </div>
              </div>
            </div>

            <div class="row mb-4">
              <div class="col-12">
                <label class="form-label fw-bold fs-6">Mode</label>
                <div class="d-flex justify-content-center gap-3 flex-wrap">
                  <div class="mode-toggle p-3 {% if request.form.get('mode') == 'curl' or not request.form.get('mode') %}active{% endif %}" onclick="toggleMode(event,'curl')">
                    <i class="fas fa-download me-2"></i>Curl (GET/POST)
                  </div>
                  <div class="mode-toggle p-3 {% if request.form.get('mode') == 'scrape' %}active{% endif %}" onclick="toggleMode(event,'scrape')">
                    <i class="fas fa-code me-2"></i>Scrape (CSS/Regex)
                  </div>
                </div>
              </div>
            </div>

            <div id="curlOptions" class="{% if request.form.get('mode') == 'scrape' %}d-none{% endif %}">
              <div class="row mb-3">
                <div class="col-md-6">
                  <label class="form-label">User-Agent</label>
                  <input name="user_agent" type="text" class="form-control" placeholder="Mozilla/5.0..." value="{{ request.form.get('user_agent','') }}">
                </div>
                <div class="col-md-6">
                  <label class="form-label">Timeout (s)</label>
                  <input name="timeout" type="number" class="form-control" min="1" max="60" value="{{ request.form.get('timeout','10') }}" placeholder="10">
                </div>
              </div>
              <div class="mb-3">
                <label class="form-label">Custom Headers (JSON)</label>
                <textarea name="custom_headers" class="form-control" rows="2" placeholder='{"Authorization": "Bearer token"}'>{{ request.form.get('custom_headers','') }}</textarea>
              </div>
              <div class="form-check mb-3">
                <input class="form-check-input" type="checkbox" name="post_method" {% if request.form.get('post_method') %}checked{% endif %}>
                <label class="form-check-label">POST Method</label>
              </div>
              <div class="form-check mb-3">
                <input class="form-check-input" type="checkbox" name="headers_only" {% if request.form.get('headers_only') %}checked{% endif %}>
                <label class="form-check-label">Headers Only</label>
              </div>
              <div class="mb-3" id="postData" class="{% if not request.form.get('post_method') %}d-none{% endif %}">
                <label class="form-label">POST Data (JSON)</label>
                <textarea name="post_data" class="form-control" rows="3" placeholder='{"key": "value"}'>{{ request.form.get('post_data','') }}</textarea>
              </div>
            </div>

            <div id="scrapeOptions" class="{% if request.form.get('mode') != 'scrape' %}d-none{% endif %}">
              <div class="mb-3">
                <label class="form-label">CSS Selector(s)</label>
                <div class="input-group">
                  <span class="input-group-text"><i class="fas fa-tag"></i></span>
                  <input name="selectors" type="text" class="form-control" placeholder="e.g. .title, .date" value="{{ request.form.get('selectors','') }}">
                </div>
              </div>
              <div class="mb-3">
                <label class="form-label">Regex Pattern (optional)</label>
                <div class="input-group">
                  <span class="input-group-text"><i class="fas fa-regex"></i></span>
                  <input name="regex_pattern" type="text" class="form-control" placeholder="e.g. \\d{4}-\\d{2}-\\d{2}" value="{{ request.form.get('regex_pattern','') }}">
                </div>
              </div>
              <div class="row">
                <div class="col-md-6">
                  <label class="form-label">User-Agent</label>
                  <input name="user_agent" type="text" class="form-control" placeholder="Mozilla/5.0..." value="{{ request.form.get('user_agent','') }}">
                </div>
                <div class="col-md-6">
                  <label class="form-label">Timeout (s)</label>
                  <input name="timeout" type="number" class="form-control" min="1" max="60" value="{{ request.form.get('timeout','10') }}" placeholder="10">
                </div>
              </div>
              <div class="form-check mt-3">
                <input class="form-check-input" type="checkbox" name="unique" {% if request.form.get('unique') %}checked{% endif %}>
                <label class="form-check-label">Unique Results Only</label>
              </div>
              <div class="form-check mt-3">
                <input class="form-check-input" type="checkbox" name="clean_data" {% if request.form.get('clean_data') %}checked{% endif %}>
                <label class="form-check-label">Clean Whitespace</label>
              </div>
            </div>

            <div class="row mb-4 mt-4">
              <div class="col-md-6">
                <label class="form-label fw-bold fs-6">Download Format</label>
                <select name="format" class="form-select">
                  <option value="csv" {% if request.form.get('format') == 'csv' %}selected{% endif %}>CSV</option>
                  <option value="json" {% if request.form.get('format') == 'json' %}selected{% endif %}>JSON</option>
                  <option value="txt" {% if request.form.get('format') == 'txt' %}selected{% endif %}>TXT</option>
                  <option value="xlsx" {% if request.form.get('format') == 'xlsx' %}selected{% endif %}>Excel</option>
                </select>
              </div>
              <div class="col-md-6 d-flex align-items-end gap-2">
                <button type="submit" class="btn btn-primary btn-modern w-100"><i class="fas fa-magic me-2"></i>Generate & Preview</button>
                <button type="button" class="btn btn-outline-light btn-modern w-100" onclick="runAutoFind()"><i class="fas fa-search me-2"></i>Auto Find Contact & Emails</button>
              </div>
            </div>
          </form>
        </div>

        {% with messages = get_flashed_messages(with_categories=true) %}
          {% if messages %}
            {% for cat, msg in messages %}
              <div class="alert alert-{{ 'danger' if cat=='error' else 'info' }} glass mb-4">{{ msg }}</div>
            {% endfor %}
          {% endif %}
        {% endwith %}

        {% if results %}
          <div class="glass p-4 mb-4">
            <h3 class="mb-3 fw-bold"><i class="fas fa-eye me-2"></i>Preview (Top 100)</h3>
            <div class="progress mb-3"><div class="progress-bar" style="width: 100%"></div></div>
            {% if table_html %}
              {{ table_html|safe }}
            {% elif raw_content %}
              <pre>{{ raw_content[:2000] }}...</pre>
            {% else %}
              <div class="alert alert-warning">No data found.</div>
            {% endif %}
            <div class="mt-4">
              <a href="{{ url_for('download') }}" class="btn btn-success btn-modern"><i class="fas fa-download me-2"></i>Download {{ request.form.get('format', 'CSV').upper() }}</a>
            </div>
          </div>

          <div class="glass p-4">
            <h5 class="mb-3 fw-bold"><i class="fas fa-info-circle me-2"></i>Metadata</h5>
            <pre>{{ metadata }}</pre>
          </div>
        {% endif %}

        {% if history %}
          <div class="glass p-4">
            <h5 class="mb-3 fw-bold"><i class="fas fa-history me-2"></i>Recent History</h5>
            <ul class="list-group">
              {% for h in history[-5:] %}
                <li class="list-group-item d-flex justify-content-between align-items-center">
                  {{ h.url }} - {{ h.mode }} ({{ h.time }})
                  <button class="btn btn-sm btn-outline-secondary btn-modern" onclick="loadHistory('{{ h.url }}', '{{ h.mode }}')">Load</button>
                </li>
              {% endfor %}
            </ul>
          </div>
        {% endif %}

        <div class="text-center text-muted mt-5">
          <small class="fs-6">
            Powered by Flask & BeautifulSoup | Enhanced 2025 <br>
            Developed by Abdul Jabbar Akhtar
          </small>
        </div>
      </div>
    </div>
  </div>

  <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
  <script>
    function toggleMode(e, mode) {
      document.getElementById('mode').value = mode;
      document.getElementById('curlOptions').classList.toggle('d-none', mode !== 'curl');
      document.getElementById('scrapeOptions').classList.toggle('d-none', mode !== 'scrape');
      document.querySelectorAll('.mode-toggle').forEach(el => el.classList.remove('active'));
      e.currentTarget.classList.add('active');
      if (mode === 'curl') document.getElementById('postData').classList.toggle('d-none', !document.querySelector('input[name="post_method"]').checked);
    }
    const postCheckbox = document.querySelector('input[name="post_method"]');
    if(postCheckbox) postCheckbox.addEventListener('change', function() {
      const pd = document.getElementById('postData'); if(pd) pd.classList.toggle('d-none', !this.checked);
    });
    function toggleDarkMode() {
      const themeInput = document.getElementById('theme');
      const currentTheme = themeInput.value;
      const newTheme = currentTheme === 'light' ? 'dark' : currentTheme === 'dark' ? 'dark-alt' : 'light';
      themeInput.value = newTheme;
      document.body.classList.remove('theme-light', 'theme-dark', 'theme-dark-alt');
      document.body.classList.add('theme-' + newTheme);
      document.documentElement.setAttribute('data-bs-theme', newTheme === 'dark-alt' ? 'light' : newTheme);
    }
    function loadHistory(url, mode) {
      document.querySelector('input[name="url"]').value = url;
      document.getElementById('mode').value = mode;
      toggleMode({currentTarget: document.querySelector('.mode-toggle.' + (mode === 'curl' ? 'active' : ''))}, mode);
      document.getElementById('scrapeForm').submit();
    }
    function runAutoFind() {
      document.getElementById('autofind').value = '1';
      document.getElementById('scrapeForm').submit();
    }
  </script>
</body>
</html>
"""

# store last results in a module-level variable; we will assign with global keyword
_LAST_RESULTS = {}

# Helpful regex presets
REGEX_PRESETS = {
    'email': r"[\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,}",
    'phone': r"\+?\d[\d\s().-]{6,}\d",
    'url': r"https?://[\w\-._~:/?#\[\]@!$&'()*+,;=]+"
}

def is_valid_url(url: str) -> bool:
    try:
        p = urlparse(url)
        return p.scheme in ("http", "https") and p.netloc
    except Exception:
        return False

def fetch_data(url: str, user_agent: str = None, timeout: int = 10, headers_only: bool = False, method: str = 'GET', custom_headers: dict = None, post_data: dict = None):
    """Fetch content; return tuple (content_text_or_bytes, content_type, headers_dict)."""
    headers = {"User-Agent": user_agent or "Mozilla/5.0 (compatible; EnhancedScraper/1.0)"}
    if custom_headers:
        headers.update(custom_headers)
    try:
        resp = requests.request(method, url, headers=headers, timeout=timeout, json=post_data, allow_redirects=True)
        resp.raise_for_status()
    except Exception:
        raise

    resp_headers = dict(resp.headers)
    if headers_only:
        return (json.dumps(resp_headers, indent=2), resp_headers.get("Content-Type", "text/plain"), resp_headers)

    try:
        text = resp.text
    except Exception:
        try:
            text = resp.content.decode('utf-8', errors='replace')
        except Exception:
            text = str(resp.content)

    return (text, resp_headers.get("Content-Type", ""), resp_headers)

def clean_text(text: str) -> str:
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text.strip())

def extract_emails(text: str):
    return re.findall(REGEX_PRESETS['email'], text or '', flags=re.IGNORECASE)

def find_contact_links(soup: BeautifulSoup, base_url: str):
    """Return absolute URLs for links that likely point to contact pages."""
    candidates = []
    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        low = href.lower()
        if any(k in low for k in ('contact', 'support', 'about', 'customer-service', 'inquiry', 'contact-us', 'contactus')):
            absolute = urljoin(base_url, href)
            candidates.append((absolute, a.get_text(strip=True) or href))
    seen = set(); out = []
    for u, text in candidates:
        if u not in seen:
            seen.add(u); out.append((u, text))
    return out

@app.route("/", methods=["GET", "POST"])
def index():
    if 'history' not in session:
        session['history'] = []
    theme = request.form.get('theme', request.args.get('theme', 'light'))
    if request.method == "POST":
        return redirect(url_for("process"))
    return render_template_string(TEMPLATE, results=False, request=request, theme=theme, history=session.get('history', []))

@app.route("/process", methods=["POST"])
def process():
    global _LAST_RESULTS
    url = request.form.get("url", "").strip()
    mode = request.form.get("mode", "curl")  # Preserve mode from form
    theme = request.form.get("theme", "light")
    user_agent = request.form.get("user_agent", "").strip() or None
    try:
        timeout = int(request.form.get("timeout", 10))
    except ValueError:
        timeout = 10
    fmt = request.form.get("format", "csv")
    headers_only = bool(request.form.get("headers_only"))
    unique = bool(request.form.get("unique"))
    clean_data_flag = bool(request.form.get("clean_data"))
    post_method = bool(request.form.get("post_method"))
    custom_headers_raw = request.form.get("custom_headers", "").strip()
    post_data_raw = request.form.get("post_data", "").strip()
    selectors_raw = request.form.get("selectors", "").strip() if mode == "scrape" else ""
    regex_pattern = request.form.get("regex_pattern", "").strip() if mode == "scrape" else ""
    autofind = request.form.get('autofind', '0') == '1'

    if not is_valid_url(url):
        flash("Invalid URL.", "error")
        return redirect(url_for("index"))
    if mode == "scrape" and not selectors_raw and not regex_pattern and not autofind:
        flash("Enter CSS selectors or regex for scrape mode (or use Auto Find).", "error")
        return redirect(url_for("index"))
    if timeout < 1 or timeout > 120:
        flash("Timeout 1-120s.", "error")
        return redirect(url_for("index"))

    try:
        custom_headers = json.loads(custom_headers_raw) if custom_headers_raw else {}
    except json.JSONDecodeError:
        flash("Invalid JSON in custom headers.", "error")
        return redirect(url_for("index"))

    try:
        post_data = json.loads(post_data_raw) if post_data_raw else None
    except json.JSONDecodeError:
        flash("Invalid JSON in POST data.", "error")
        return redirect(url_for("index"))

    try:
        method = 'POST' if post_method else 'GET'

        if autofind:
            home_html, home_ctype, home_headers = fetch_data(url, user_agent, timeout, False, method, custom_headers, post_data)
            soup = BeautifulSoup(home_html, 'html.parser')
            contact_links = find_contact_links(soup, url)

            results_rows = []
            visited = set()
            emails_home = extract_emails(home_html)
            for e in emails_home:
                results_rows.append([url, 'homepage', e])
                visited.add((url, e))

            for link, text in contact_links:
                try:
                    if link in visited:
                        continue
                    page_html, _, _ = fetch_data(link, user_agent, timeout, False, method, custom_headers, post_data)
                    emails = extract_emails(page_html)
                    if not emails:
                        for a in BeautifulSoup(page_html, 'html.parser').select('a[href^="mailto:"]'):
                            mail = a.get('href').split(':', 1)[1] if ':' in a.get('href') else a.get('href')
                            if mail:
                                emails.append(mail)
                    for e in emails:
                        if (link, e) not in visited:
                            results_rows.append([link, text, e])
                            visited.add((link, e))
                except Exception:
                    continue

            df = pd.DataFrame(results_rows, columns=['source_url', 'link_text', 'email']) if results_rows else pd.DataFrame(columns=['source_url', 'link_text', 'email'])
            table_html = df.head(200).to_html(classes="table table-striped table-hover", index=False, escape=False)
            results = {"table_html": table_html, "mode": "autofind", "rows": df.values.tolist(), "columns": df.columns.tolist()}
            metadata = f"AutoFind run: {datetime.now().isoformat()}\nHome: {url}\nContact candidates: {len(contact_links)}\nEmails found: {len(df)}"

        elif mode == "curl":
            content, ctype, headers = fetch_data(url, user_agent, timeout, headers_only, method, custom_headers, post_data)
            raw_preview = None
            try:
                parsed = json.loads(content)
                raw_preview = json.dumps(parsed, indent=2)
            except Exception:
                raw_preview = content if isinstance(content, str) else str(content)

            results = {"raw_content": raw_preview if len(str(raw_preview)) < 10000 else str(raw_preview)[:10000] + '...', "mode": "curl", "headers": headers}
            metadata = f"Fetched ({method}): {datetime.now().isoformat()}\nContent-Type: {ctype}\nLength: {len(raw_preview)}"

        else:  # scrape
            html, ctype, headers = fetch_data(url, user_agent, timeout, False, method, custom_headers, post_data)
            soup = BeautifulSoup(html, "html.parser")
            selectors = [s.strip() for s in selectors_raw.split(",") if s.strip()] if selectors_raw else []

            lists_by_selector = []
            max_len = 0

            if selectors:
                for sel in selectors:
                    els = soup.select(sel)
                    texts = []
                    for el in els:
                        txt = el.get_text(separator=' ', strip=True) or next((el.get(a) for a in ["alt", "title", "src", "href"] if el.get(a)), "")
                        if clean_data_flag:
                            txt = clean_text(txt)
                        if regex_pattern:
                            matches = re.findall(regex_pattern, txt, flags=re.DOTALL)
                            if matches:
                                txt = ', '.join(matches)
                            else:
                                txt = ''
                        texts.append(txt)
                    lists_by_selector.append(texts[:100])
                    if len(texts) > max_len:
                        max_len = len(texts)
            else:
                matches = re.findall(regex_pattern, html, flags=re.DOTALL) if regex_pattern else []
                normalized = [m[0] if isinstance(m, tuple) else m for m in matches][:100]
                lists_by_selector = [[match] for match in normalized]
                max_len = len(lists_by_selector)

            rows = []
            for i in range(max_len):
                row = [lst[i] if i < len(lst) else "" for lst in lists_by_selector]
                rows.append(row)

            if unique:
                rows = list({tuple(r): r for r in rows}.values())

            columns = [f"{s[:20]}..." if len(s) > 20 else s for s in selectors] or ["regex_match"]
            if rows and len(columns) != len(rows[0]):
                columns = [f"Column_{i+1}" for i in range(len(rows[0]))]

            df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
            table_html = df.head(100).to_html(classes="table table-striped table-hover", index=False, escape=False)
            results = {"table_html": table_html, "mode": "scrape", "rows": rows, "columns": columns}
            metadata = f"Scraped ({method}): {datetime.now().isoformat()}\nSelectors/Regex: {selectors_raw or regex_pattern}\nRows: {len(rows)}\nUnique: {unique}\nClean: {clean_data_flag}"

        _LAST_RESULTS = {"results": results, "url": url, "mode": mode, "format": fmt, "metadata": metadata}
        session['history'].append({"url": url, "mode": mode, "time": datetime.now().strftime("%Y-%m-%d %H:%M")})
        session.modified = True
        return render_template_string(TEMPLATE, results=True, table_html=results.get("table_html"), raw_content=results.get("raw_content"), metadata=metadata, request=request, theme=theme, history=session.get('history', []))

    except requests.exceptions.RequestException as rexc:
        traceback.print_exc()
        flash(f"Request error: {str(rexc)}", "error")
        return redirect(url_for("index"))
    except Exception as e:
        traceback.print_exc()
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("index"))

@app.route("/download")
def download():
    global _LAST_RESULTS
    data = _LAST_RESULTS
    if not data:
        flash("No data.", "error")
        return redirect(url_for("index"))

    results = data["results"]
    fmt = data["format"]
    mode = data["mode"]

    if mode == 'autofind':
        df = pd.DataFrame(results.get('rows', []), columns=results.get('columns', ['source_url', 'link_text', 'email'])) \
             if results.get('rows') else pd.DataFrame(columns=results.get('columns', ['source_url', 'link_text', 'email']))
        if fmt == 'csv':
            mem = io.StringIO()
            df.to_csv(mem, index=False)
            return send_file(io.BytesIO(mem.getvalue().encode()), as_attachment=True, download_name='autofind_emails.csv', mimetype='text/csv')
        if fmt == 'json':
            return send_file(io.BytesIO(df.to_json(orient='records').encode()), as_attachment=True, download_name='autofind_emails.json', mimetype='application/json')
        if fmt == 'xlsx':
            mem = io.BytesIO()
            with pd.ExcelWriter(mem, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            mem.seek(0)
            return send_file(mem, as_attachment=True, download_name='autofind_emails.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        mem = io.StringIO()
        df.to_string(mem, index=False)
        return send_file(io.BytesIO(mem.getvalue().encode()), as_attachment=True, download_name='autofind_emails.txt', mimetype='text/plain')

    if mode == 'curl':
        content = results.get("raw_content", "")
        if fmt == "txt":
            return send_file(io.BytesIO(str(content).encode()), as_attachment=True, download_name=f"curl_{urlparse(data['url']).netloc}.txt", mimetype='text/plain')
        elif fmt == "json":
            try:
                parsed = json.loads(content)
                return send_file(io.BytesIO(json.dumps(parsed, indent=2).encode()), as_attachment=True, download_name='curl.json', mimetype='application/json')
            except Exception:
                return send_file(io.BytesIO(json.dumps({"content": content}).encode()), as_attachment=True, download_name='curl.json', mimetype='application/json')
        elif fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Content"
            ws['A2'] = str(content)
            mem = io.BytesIO()
            wb.save(mem)
            mem.seek(0)
            return send_file(mem, as_attachment=True, download_name='curl.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:  # csv
            mem = io.StringIO()
            mem.write('content\n')
            mem.write(str(content).replace('\n', '\\n'))
            return send_file(io.BytesIO(mem.getvalue().encode()), as_attachment=True, download_name='curl.csv', mimetype='text/csv')

    df = pd.DataFrame(results.get("rows", []), columns=results.get("columns", [])) if results.get('rows') else pd.DataFrame(columns=results.get('columns', []))
    if fmt == "csv":
        mem = io.StringIO()
        df.to_csv(mem, index=False)
        return send_file(io.BytesIO(mem.getvalue().encode()), as_attachment=True, download_name="scraped.csv", mimetype="text/csv")
    elif fmt == "json":
        return send_file(io.BytesIO(df.to_json(orient="records").encode()), as_attachment=True, download_name="scraped.json", mimetype="application/json")
    elif fmt == "xlsx":
        mem = io.BytesIO()
        with pd.ExcelWriter(mem, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        mem.seek(0)
        return send_file(mem, as_attachment=True, download_name="scraped.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:  # txt
        mem = io.StringIO()
        df.to_string(mem, index=False)
        return send_file(io.BytesIO(mem.getvalue().encode()), as_attachment=True, download_name="scraped.txt", mimetype='text/plain')

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
